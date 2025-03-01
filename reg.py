import tkinter as tk
from tkinter import messagebox
from tkinter import Label, Frame, Canvas
from PIL import Image, ImageTk
import subprocess
import sys
import openpyxl
import openpyxl
from tkinter import messagebox

def register_user():
    """Registers the user and saves their data in an Excel file."""
    # Get user inputs
    username = entry1.get()
    email = entry2.get()
    password = entry3.get()
    confirm_password = entry4.get()

    # Validate inputs
    if not username or not email or not password:
        messagebox.showerror("Error", "All fields are required!")
        return

    if password != confirm_password:
        messagebox.showerror("Error", "Passwords do not match!")
        return

    # Store registration details in Excel file
    try:
        # Open or create the Excel workbook
        try:
            wb = openpyxl.load_workbook('registration_data.xlsx')
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            # Create headers for the columns if the file is new
            sheet['A1'] = 'Username'
            sheet['B1'] = 'Email'
            sheet['C1'] = 'Password'

        # Select the active sheet
        sheet = wb.active

        # Find the next available row
        next_row = sheet.max_row + 1

        # Write the registration details to the next available row
        sheet[f'A{next_row}'] = username
        sheet[f'B{next_row}'] = email
        sheet[f'C{next_row}'] = password

        # Save the workbook
        wb.save('registration_data.xlsx')

        # Show success message
        messagebox.showinfo("Success", "Registration Successful!")
        clear_form()  # Clear the form fields after registration

    except Exception as e:
        # Show error message if an issue occurs
        messagebox.showerror("Error", f"An error occurred: {e}")


def clear_form():
    """Clears all form fields."""
    entry1.delete(0, tk.END)
    entry2.delete(0, tk.END)
    entry3.delete(0, tk.END)
    entry4.delete(0, tk.END)


def open_main_app():
    root.destroy()
    subprocess.Popen([sys.executable, "login.py"])
# Create the main Tkinter window
root = tk.Tk()
root.title("Vocal Script - Signup Page")
root.geometry("1366x768")

# Load and set the background image
bg_image = Image.open("bg1.jpg")  # Replace with your image path
bg_image = bg_image.resize((1550, 780), Image.Resampling.LANCZOS)
bg_photo = ImageTk.PhotoImage(bg_image)

# Create a Canvas and add the background image
canvas = Canvas(root, width=1550, height=780)
canvas.pack(fill="both", expand=True)
canvas.create_image(0, 0, anchor="nw", image=bg_photo)

# Prevent white space by resizing dynamically
def resize_image(event):
    new_width = event.width
    new_height = event.height
    resized_image = bg_image.resize((new_width, new_height), Image.Resampling.LANCZOS)
    updated_bg_photo = ImageTk.PhotoImage(resized_image)
    canvas.create_image(0, 0, anchor="nw", image=updated_bg_photo)
    canvas.image = updated_bg_photo  # Keep a reference to avoid garbage collection


# Bind resizing events
root.bind("<Configure>", resize_image)

# Header Frame
header_frame = Frame(root, bg="#1f1c2c", height=150)
header_frame.place(x=0, y=0, relwidth=1)

# Load and resize the logo
original_logo = Image.open("hlogo.png")
resized_logo = original_logo.resize((130, 130), Image.Resampling.LANCZOS)
logo_image = ImageTk.PhotoImage(resized_logo)

# Place logo in the header frame
logo_label = Label(header_frame, image=logo_image, bg="#2c2738")
logo_label.place(x=20, y=10)

# Header Label
header_label = tk.Label(
    header_frame,
    text=" Vocal Script",
    font=("Helvetica", 32, "bold"),
    bg="#1f1c2c",
    fg="#ffffff",
    pady=10,
)
header_label.place(x=180, y=37)

# Caption Label
caption_label = Label(
    header_frame,
    text="  Say It Once, Keep It Forever!",
    font=("Helvetica", 15, "italic"),
    fg="white",
    bg="#1f1c2c",
)
caption_label.place(x=188, y=93)

# Form Frame
form_frame = tk.Frame(root, bg="#1f1c2c", padx=20, pady=20, relief="raised", borderwidth=3)
form_frame.place(relx=0.5, rely=0.5, anchor="center")

heading_label = tk.Label(
    form_frame, text="Signup", font=("Helvetica", 20, "bold"), bg="#1f1c2c", fg="white"
)
heading_label.grid(row=0, column=0, columnspan=2, pady=10)

# Username
label1 = tk.Label(form_frame, text="Username:", font=("Arial", 16, "bold"), bg="#1f1c2c", fg="white")
label1.grid(row=1, column=0, padx=10, pady=10, sticky="w")
entry1 = tk.Entry(form_frame, font=("Arial", 16), width=25)
entry1.grid(row=1, column=1, padx=10, pady=10)

# Email
label2 = tk.Label(form_frame, text="Email:", font=("Arial", 16, "bold"), bg="#1f1c2c", fg="white")
label2.grid(row=2, column=0, padx=10, pady=10, sticky="w")
entry2 = tk.Entry(form_frame, font=("Arial", 16), width=25)
entry2.grid(row=2, column=1, padx=10, pady=10)

# Password
label3 = tk.Label(form_frame, text="Password:", font=("Arial", 16, "bold"), bg="#1f1c2c", fg="white")
label3.grid(row=3, column=0, padx=10, pady=10, sticky="w")
entry3 = tk.Entry(form_frame, font=("Arial", 16), width=25, show="*")
entry3.grid(row=3, column=1, padx=10, pady=10)

# Confirm Password
label4 = tk.Label(form_frame, text="Confirm Password:", font=("Arial", 16, "bold"), bg="#1f1c2c", fg="white")
label4.grid(row=4, column=0, padx=10, pady=10, sticky="w")
entry4 = tk.Entry(form_frame, font=("Arial", 16), width=25, show="*")
entry4.grid(row=4, column=1, padx=10, pady=10)

# Buttons
button_frame = tk.Frame(form_frame, bg="#1f1c2c")
button_frame.grid(row=5, column=0, columnspan=2, pady=20)

button_register = tk.Button(
    button_frame,
    text="Register",
    font=("Arial", 12, "bold"),
    bg="green",
    fg="white",
    width=12,
    command=register_user,
)
button_register.grid(row=0, column=0, padx=10)

button_clear = tk.Button(
    button_frame,
    text="Clear",
    font=("Arial", 12, "bold"),
    bg="orange",
    fg="white",
    width=12,
    command=clear_form,
)
button_clear.grid(row=0, column=1, padx=10)

button_exit = tk.Button(
    button_frame,
    text="Login",
    font=("Arial", 12, "bold"),
    bg="#1976D2",
    fg="white",
    width=12,
    command=open_main_app,
)
button_exit.grid(row=0, column=2, padx=10)



# Run the application
root.mainloop()
